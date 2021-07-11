import openpyxl

path="E:\\pythonProject\\pawanOnlineSeleniumPython\\xlData.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.active

rows = sheet.max_row
cols = sheet.max_column

print("Number of row: ",rows)
print("Number of columns: ",cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end='   ')
    print()